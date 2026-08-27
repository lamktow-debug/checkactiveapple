import asyncio
import os
import unittest
from tempfile import NamedTemporaryFile
from pathlib import Path
import sys
import json
import types
from datetime import date
from unittest.mock import AsyncMock, MagicMock, patch
from openpyxl import load_workbook

fake_playwright = types.ModuleType("playwright")
fake_playwright_async_api = types.ModuleType("playwright.async_api")


async def _fake_async_playwright():
    return None


fake_playwright_async_api.async_playwright = _fake_async_playwright
fake_playwright.async_api = fake_playwright_async_api
sys.modules.setdefault("playwright", fake_playwright)
sys.modules.setdefault("playwright.async_api", fake_playwright_async_api)

fake_playwright_stealth = types.ModuleType("playwright_stealth")


class _FakeStealth:
    def use_async(self, playwright_context):
        return playwright_context


fake_playwright_stealth.Stealth = _FakeStealth
sys.modules.setdefault("playwright_stealth", fake_playwright_stealth)

from check_active_v2 import (
    CAPTCHA_2CAPTCHA_API_KEY,
    BlockThrottle,
    BlockedError,
    CaptchaBusyError,
    CaptchaServiceError,
    CaptchaSolution,
    CAPTCHA_2CAPTCHA_CREATE_TASK_URL,
    CAPTCHA_2CAPTCHA_FIRST_POLL_DELAY_SECONDS,
    CAPTCHA_2CAPTCHA_MAX_POLLS,
    contains_captcha_error_text,
    ensure_serial_filled,
    wait_for_result_payload,
    CAPTCHA_RELOAD_POLL_MS,
    CAPTCHA_RELOAD_MAX_WAIT_MS,
    RESULT_TEXT_KEYS,
    reload_captcha,
    save_result_screenshot,
    CAPTCHA_2CAPTCHA_POLL_INTERVAL_SECONDS,
    CAPTCHA_2CAPTCHA_RESULT_URL,
    CAPTCHA_2CAPTCHA_REPORT_BAD_URL,
    post_2captcha_json,
    SCREENSHOT_DIR,
    build_screenshot_path,
    check_serial,
    create_stealth_playwright_context,
    determine_purchase_date,
    export_inactive_to_excel,
    extract_visible_text,
    fill_serial_number,
    has_result_page_signal,
    contains_captcha_error_text,
    ensure_serial_filled,
    build_inactive_excel_path,
    get_run_folder_name,
    is_captcha_wait_timeout_error,
    is_serial_input_timeout_error,
    normalize_captcha_code,
    MAX_AUTO_RETRIES,
    prompt_run_settings,
    read_result_texts,
    should_capture_screenshot,
    classify_2captcha_error,
    is_valid_serial,
    load_proxies,
    load_serials,
    needs_check,
    normalize_serial,
    open_check_page,
    BLOCKED_RESOURCE_TYPES,
    CAPTCHA_MAX_LENGTH,
    CAPTCHA_MIN_LENGTH,
    ProxyFailure,
    block_heavy_assets,
    should_block_request,
    is_page_closed_error,
    pick_proxy,
    clear_site_state,
    parse_args,
    parse_proxy_line,
    proxy_for_index,
    solve_captcha_task,
    submit_captcha_code,
    wait_for_result_payload,
)


class ScreenshotPathTests(unittest.TestCase):
    def test_build_screenshot_path_uses_serial_name_for_active_devices(self):
        serial = "ABC123XYZ"
        run_date = date(2026, 3, 30)

        screenshot_path = build_screenshot_path(serial, "01/03/2026", run_date)

        self.assertEqual(
            screenshot_path,
            SCREENSHOT_DIR / "300326" / "Đã active" / "ABC123XYZ.png",
        )
        self.assertTrue((SCREENSHOT_DIR / "300326" / "Đã active").exists())

    def test_build_screenshot_path_uses_inactive_folder_and_sanitizes_name(self):
        serial = 'AB:C/12*3?"<>|'
        run_date = date(2026, 3, 30)

        screenshot_path = build_screenshot_path(serial, "Chưa active", run_date)

        self.assertEqual(
            screenshot_path,
            SCREENSHOT_DIR / "300326" / "Chưa active" / "AB_C_12_3_____.png",
        )
        self.assertIsInstance(screenshot_path, Path)
        self.assertTrue((SCREENSHOT_DIR / "300326" / "Chưa active").exists())

    def test_get_run_folder_name_formats_date_as_ddmmyy(self):
        self.assertEqual(get_run_folder_name(date(2026, 3, 30)), "300326")

    def test_get_run_folder_name_returns_custom_name_when_provided(self):
        self.assertEqual(get_run_folder_name(date(2026, 3, 30), "batch_april"), "batch_april")

    def test_should_capture_screenshot_accepts_y_and_n_inputs(self):
        self.assertTrue(should_capture_screenshot("y"))
        self.assertTrue(should_capture_screenshot("Y"))
        self.assertFalse(should_capture_screenshot("n"))
        self.assertFalse(should_capture_screenshot("N"))

    def test_is_captcha_wait_timeout_error_detects_playwright_timeout_message(self):
        error_message = (
            'Locator.wait_for: Timeout 10000ms exceeded.\n'
            'Call log:\n'
            '  - waiting for locator("img.captcha-image, img[alt=\\"captcha\\"]").first to be visible'
        )
        self.assertTrue(is_captcha_wait_timeout_error(RuntimeError(error_message)))
        self.assertFalse(is_captcha_wait_timeout_error(RuntimeError("some other error")))

    def test_is_serial_input_timeout_error_detects_fill_timeout_message(self):
        error_message = (
            'Page.fill: Timeout 30000ms exceeded.\n'
            'Call log:\n'
            '  - waiting for locator("#serial-number-input")'
        )
        self.assertTrue(is_serial_input_timeout_error(RuntimeError(error_message)))
        self.assertFalse(is_serial_input_timeout_error(RuntimeError("some other error")))

    def test_normalize_captcha_code_keeps_only_ascii_letters_and_digits(self):
        self.assertEqual(normalize_captcha_code("abçĐ-12@*z"), "AB12Z")

    def test_solve_captcha_task_sends_base64_image_to_2captcha_and_polls_result(self):
        requests = []

        class FakeResponse:
            def __init__(self, payload):
                self._payload = payload

            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return False

            def read(self):
                return json.dumps(self._payload).encode("utf-8")

        def fake_urlopen(request, timeout):
            requests.append((request.full_url, json.loads(request.data.decode("utf-8")), timeout))
            if request.full_url == CAPTCHA_2CAPTCHA_CREATE_TASK_URL:
                return FakeResponse({"errorId": 0, "taskId": 12345})
            return FakeResponse({"errorId": 0, "status": "ready", "solution": {"text": "ab-12"}})

        with (
            patch("check_active_v2.urlopen", side_effect=fake_urlopen),
            patch("check_active_v2.asyncio.sleep", new=AsyncMock()),
        ):
            solution = asyncio.run(solve_captcha_task("data:image/png;base64,QUJDRA=="))

        self.assertEqual(solution.code, "AB12")
        self.assertEqual(solution.task_id, 12345)
        self.assertEqual(requests[0][0], CAPTCHA_2CAPTCHA_CREATE_TASK_URL)
        self.assertEqual(requests[0][1]["clientKey"], CAPTCHA_2CAPTCHA_API_KEY)
        self.assertEqual(requests[0][1]["task"]["type"], "ImageToTextTask")
        self.assertEqual(requests[0][1]["task"]["body"], "QUJDRA==")
        self.assertEqual(requests[0][1]["task"]["minLength"], 4)
        self.assertEqual(requests[0][1]["task"]["maxLength"], 4)
        self.assertEqual(requests[1][0], CAPTCHA_2CAPTCHA_RESULT_URL)
        self.assertEqual(requests[1][1]["taskId"], 12345)

    def _fake_2captcha_response(self, payload):
        class FakeResponse:
            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return False

            def read(self):
                return json.dumps(payload).encode("utf-8")

        return FakeResponse()

    def test_solve_captcha_task_raises_for_fatal_2captcha_error(self):
        response = self._fake_2captcha_response(
            {"errorId": 1, "errorCode": "ERROR_ZERO_BALANCE"}
        )
        with patch("check_active_v2.urlopen", return_value=response):
            with self.assertRaises(CaptchaServiceError):
                asyncio.run(solve_captcha_task("QUJDRA=="))

    def test_solve_captcha_task_raises_busy_error_when_no_slot_available(self):
        response = self._fake_2captcha_response(
            {"errorId": 1, "errorCode": "ERROR_NO_SLOT_AVAILABLE"}
        )
        with patch("check_active_v2.urlopen", return_value=response):
            with self.assertRaises(CaptchaBusyError):
                asyncio.run(solve_captcha_task("QUJDRA=="))

    def test_solve_captcha_task_returns_empty_for_ordinary_error(self):
        response = self._fake_2captcha_response(
            {"errorId": 1, "errorCode": "ERROR_IMAGE_TYPE_NOT_SUPPORTED"}
        )
        with patch("check_active_v2.urlopen", return_value=response):
            solution = asyncio.run(solve_captcha_task("QUJDRA=="))

        self.assertEqual(solution.code, "")

    def test_classify_2captcha_error_returns_none_when_no_error(self):
        self.assertIsNone(classify_2captcha_error({"errorId": 0}))

    def test_create_stealth_playwright_context_wraps_async_playwright_with_stealth(self):
        class FakeStealth:
            created_with = None

            def use_async(self, playwright_context):
                FakeStealth.created_with = playwright_context
                return "stealth-context"

        with (
            patch("check_active_v2.Stealth", new=FakeStealth),
            patch("check_active_v2.async_playwright", new=lambda: "playwright-context"),
        ):
            context = create_stealth_playwright_context()

        self.assertEqual(context, "stealth-context")
        self.assertEqual(FakeStealth.created_with, "playwright-context")

    def test_determine_purchase_date_detects_inactive_from_header_without_device_title(self):
        purchase_date = determine_purchase_date(
            device_title="",
            header_text="MacBook Air\nThiết bị chưa được kích hoạt",
            purchase_text="",
            notification_text="",
        )
        self.assertEqual(purchase_date, "Chưa active")

    def test_determine_purchase_date_detects_inactive_from_notification_heading(self):
        purchase_date = determine_purchase_date(
            device_title="iPhone 17 Pro Max",
            header_text="",
            purchase_text="",
            notification_text="Thiết bị chưa được kích hoạt",
            body_text="",
        )
        self.assertEqual(purchase_date, "Chưa active")

    def test_has_result_page_signal_requires_real_result_content(self):
        self.assertFalse(
            has_result_page_signal(
                device_title="",
                header_text="",
                purchase_text="",
                notification_text="",
            )
        )
        self.assertTrue(
            has_result_page_signal(
                device_title="iPhone 17",
                header_text="Số Sê-ri: D4942403C0",
                purchase_text="",
                notification_text="",
            )
        )
        self.assertTrue(
            has_result_page_signal(
                device_title="",
                header_text="",
                purchase_text="",
                notification_text="Thiết bị chưa được kích hoạt",
            )
        )

    def test_detects_captcha_mismatch_text(self):
        self.assertTrue(
            contains_captcha_error_text(
                'Rất tiếc. Mã bạn đã nhập không khớp với hình ảnh. '
                'Nếu bạn muốn một hình ảnh khác, chọn "Làm mới mã".'
            )
        )
        # "Nhap ma trong anh" la NHAN cua o nhap, khong phai loi. Coi no la loi
        # thi moi lan gui ma DUNG cung bi ket luan sai ngay khi trang chua kip
        # chuyen. Cau bao loi that da chua "Ma ban da nhap khong khop..." roi.
        self.assertFalse(contains_captcha_error_text("Nhập mã trong ảnh."))
        self.assertFalse(contains_captcha_error_text("Số sê-ri bạn đã nhập không hợp lệ."))

    def test_wait_for_result_payload_returns_immediately_on_captcha_error(self):
        class FakePage:
            wait_calls = 0

            async def wait_for_timeout(self, _ms):
                self.wait_calls += 1

        page = FakePage()
        payload = {
            "device_title": "",
            "purchase_text": "",
            "notification_text": "",
            "header_text": "",
            "body_text": "",
            "heading_text": "",
            "error_text": 'Rất tiếc. Mã bạn đã nhập không khớp với hình ảnh.',
        }

        with patch("check_active_v2.read_result_texts", new=AsyncMock(return_value=payload)):
            result = asyncio.run(wait_for_result_payload(page))

        self.assertTrue(result["captcha_error"])
        self.assertIsNone(result["purchase_date"])
        self.assertEqual(page.wait_calls, 0)

    def test_determine_purchase_date_uses_purchase_text_when_result_page_is_visible(self):
        purchase_date = determine_purchase_date(
            device_title="iPhone 15",
            header_text="iPhone 15\nĐã mua 3 tháng 4, 2026",
            purchase_text="Đã mua 3 tháng 4, 2026",
            notification_text="",
        )
        self.assertEqual(purchase_date, "03/04/2026")

    def test_determine_purchase_date_parses_apple_vietnamese_purchase_format(self):
        purchase_date = determine_purchase_date(
            device_title="iPhone 15",
            header_text="iPhone 15\nĐã mua 17 tháng 9, 2025",
            purchase_text="Đã mua 17 tháng 9, 2025",
            notification_text="",
            body_text="",
        )
        self.assertEqual(purchase_date, "17/09/2025")

    def test_determine_purchase_date_returns_unverified_for_unverified_purchase_message(self):
        purchase_date = determine_purchase_date(
            device_title="iPhone 17",
            header_text="iPhone 17\nSố Sê-ri: JYRH91FWH5\nNgày mua chưa được xác thực",
            purchase_text="",
            notification_text="",
            body_text="",
        )
        self.assertEqual(purchase_date, "Chưa xác thực")

    def test_determine_purchase_date_returns_unverified_when_message_is_only_in_body_text(self):
        purchase_date = determine_purchase_date(
            device_title="iPhone 17",
            header_text="iPhone 17\nSố Sê-ri: FP2F7VHV29",
            purchase_text="",
            notification_text="",
            body_text="Thông tin bảo hành\nNgày mua chưa được xác thực",
        )
        self.assertEqual(purchase_date, "Chưa xác thực")

    def test_determine_purchase_date_returns_invalid_for_invalid_serial_message(self):
        purchase_date = determine_purchase_date(
            device_title="",
            header_text="",
            purchase_text="",
            notification_text="",
            body_text="",
            heading_text="Số sê-ri bạn đã nhập không hợp lệ. Vui lòng thử lại.",
        )
        self.assertEqual(purchase_date, "serial ko hợp lệ")

    def test_determine_purchase_date_ignores_invalid_serial_text_when_only_in_body(self):
        purchase_date = determine_purchase_date(
            device_title="",
            header_text="",
            purchase_text="",
            notification_text="",
            body_text="Số sê-ri bạn đã nhập không hợp lệ. Vui lòng thử lại.",
        )
        self.assertIsNone(purchase_date)

    def test_determine_purchase_date_does_not_turn_serial_digits_into_fake_date(self):
        purchase_date = determine_purchase_date(
            device_title="iPhone 17",
            header_text="iPhone 17 Số Sê-ri: JYRH91FWH5 Ngày mua chưa được xác thực",
            purchase_text="",
            notification_text="",
            body_text="",
        )
        self.assertEqual(purchase_date, "Chưa xác thực")

    def test_determine_purchase_date_uses_body_text_fallback_when_selectors_miss(self):
        purchase_date = determine_purchase_date(
            device_title="",
            header_text="",
            purchase_text="",
            notification_text="",
            body_text="iPhone 15\nThông tin bảo hành\nĐã mua 3 tháng 4, 2026",
        )
        self.assertEqual(purchase_date, "03/04/2026")

    def test_determine_purchase_date_detects_single_apple_purchase_line_in_header_text(self):
        purchase_date = determine_purchase_date(
            device_title="iPhone 15",
            header_text="iPhone 15\nĐã mua 4 tháng 3, 2026",
            purchase_text="",
            notification_text="",
            body_text="",
        )
        self.assertEqual(purchase_date, "04/03/2026")

    def test_determine_purchase_date_detects_single_apple_purchase_line_in_body_text(self):
        purchase_date = determine_purchase_date(
            device_title="iPhone 15",
            header_text="",
            purchase_text="",
            notification_text="",
            body_text="iPhone 15\nThông tin bảo hành\nĐã mua 4 tháng 3, 2026",
        )
        self.assertEqual(purchase_date, "04/03/2026")

    def test_extract_visible_text_strips_text_content(self):
        class FakeLocator:
            async def count(self):
                return 1

            async def text_content(self):
                return "  Đã mua 13 tháng 1, 2026  "

        text = asyncio.run(extract_visible_text(FakeLocator()))
        self.assertEqual(text, "Đã mua 13 tháng 1, 2026")

    def test_read_result_texts_reads_body_text_for_unverified_cases(self):
        class FakeLocator:
            def __init__(self, text):
                self._text = text

            @property
            def first(self):
                return self

            async def count(self):
                return 1 if self._text is not None else 0

            async def is_visible(self):
                return bool(self._text)

            async def text_content(self):
                return self._text

        class FakePage:
            def locator(self, selector):
                mapping = {
                    "#device-header-title": FakeLocator("iPhone 17"),
                    "p.device-header-purchase": FakeLocator(""),
                    'h2.notification-heading, [data-testid="notification-heading"]': FakeLocator(""),
                    ".device-header-wrapper": FakeLocator("iPhone 17\nSố Sê-ri: FP2F7VHV29"),
                    "body": FakeLocator("Thông tin bảo hành\nNgày mua chưa được xác thực"),
                    "h1": FakeLocator(""),
                    "div.err-msg-container, .err-msg": FakeLocator(""),
                }
                return mapping[selector]

        payload = asyncio.run(read_result_texts(FakePage()))
        self.assertEqual(payload["body_text"], "Thông tin bảo hành\nNgày mua chưa được xác thực")

    def test_prompt_run_settings_retries_until_valid_screenshot_input(self):
        with patch("builtins.input", side_effect=["x", "maybe", "y", ""]):
            settings = prompt_run_settings(date(2026, 3, 30))

        self.assertEqual(settings["folder_name"], "300326")
        self.assertTrue(settings["capture_screenshot"])

    def test_prompt_run_settings_skips_folder_prompt_when_screenshot_disabled(self):
        with patch("builtins.input", side_effect=["n"]) as mock_input:
            settings = prompt_run_settings(date(2026, 3, 30))

        self.assertIsNone(settings["folder_name"])
        self.assertFalse(settings["capture_screenshot"])
        self.assertEqual(mock_input.call_count, 1)

    def test_export_inactive_to_excel_writes_file_in_inactive_folder(self):
        run_date = date(2026, 3, 30)
        rows = [
            ["SN_ACTIVE", "01/03/2026"],
            ["SN_INACTIVE_1", "Chưa active"],
            ["SN_INACTIVE_2", "Chưa active"],
        ]

        excel_path = export_inactive_to_excel(rows, run_date)

        self.assertEqual(
            excel_path,
            SCREENSHOT_DIR / "300326" / "Chưa active" / "Chưa active.xlsx",
        )
        self.assertEqual(excel_path, build_inactive_excel_path(run_date))
        self.assertTrue(excel_path.exists())

        workbook = load_workbook(excel_path)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))

        self.assertEqual(values[0], ("Serial", "Tên Máy"))
        self.assertEqual(values[1], ("SN_INACTIVE_1", "Chưa active"))
        self.assertEqual(values[2], ("SN_INACTIVE_2", "Chưa active"))
        self.assertEqual(len(values), 3)


class CheckSerialScreenshotBehaviorTests(unittest.TestCase):
    def test_check_serial_submits_captcha_before_trusting_serial_validation(self):
        class FakeLocator:
            def __init__(self, text="", visible=False):
                self._text = text
                self._visible = visible
                self.first = self

            async def count(self):
                return 1 if self._text is not None else 0

            async def is_visible(self):
                return self._visible

            async def text_content(self):
                return self._text

            async def wait_for(self, *_args, **_kwargs):
                return None

            async def clear(self):
                return None

            async def press_sequentially(self, *_args, **_kwargs):
                return None

            async def evaluate(self, *_args, **_kwargs):
                return None

        class FakePage:
            keyboard = AsyncMock()

            def locator(self, selector):
                if selector == 'div.err-msg-container, .err-msg':
                    return FakeLocator("Vui lòng nhập số sê-ri hợp lệ.", True)
                return FakeLocator("", False)

        fake_page = FakePage()

        with (
            patch("check_active_v2.open_check_page", new=AsyncMock(return_value=True)),
            patch("check_active_v2.get_captcha_image_base64", new=AsyncMock(return_value="QUJDRA==")),
            patch("check_active_v2.solve_captcha_task", new=AsyncMock(return_value=CaptchaSolution("ABCD", 1))) as mock_solve,
            patch("check_active_v2.wait_for_result_payload",
                  new=AsyncMock(return_value={"purchase_date": "serial ko hợp lệ"})),
        ):
            result = asyncio.run(check_serial(fake_page, "D96XVQJRD8", capture_screenshot=False))

        self.assertEqual(result, ["D96XVQJRD8", "serial ko hợp lệ"])
        mock_solve.assert_awaited_once()

    def test_check_serial_skips_screenshot_for_inactive_result(self):
        fake_locator = AsyncMock()
        fake_locator.first = fake_locator

        class FakePage:
            def __init__(self, locator, keyboard):
                self._locator = locator
                self.keyboard = keyboard

            def locator(self, _selector):
                return self._locator

        fake_page = FakePage(fake_locator, AsyncMock())

        with (
            patch("check_active_v2.open_check_page", new=AsyncMock(return_value=True)),
            patch("check_active_v2.solve_captcha_task", new=AsyncMock(return_value=CaptchaSolution("ABCD", 1))),
            patch("check_active_v2.get_captcha_image_base64", new=AsyncMock(return_value="QUJDRA==")),
            patch(
                "check_active_v2.wait_for_result_payload",
                new=AsyncMock(return_value={"purchase_date": "Chưa active"}),
            ),
            patch("check_active_v2.save_result_screenshot", new=AsyncMock()) as mock_screenshot,
        ):
            result = asyncio.run(check_serial(fake_page, "SN0000001", capture_screenshot=True, folder_name="batch"))

        self.assertEqual(result, ["SN0000001", "Chưa active"])
        mock_screenshot.assert_not_awaited()

    def test_check_serial_keeps_screenshot_for_active_result(self):
        fake_locator = AsyncMock()
        fake_locator.first = fake_locator

        class FakePage:
            def __init__(self, locator, keyboard):
                self._locator = locator
                self.keyboard = keyboard

            def locator(self, _selector):
                return self._locator

        fake_page = FakePage(fake_locator, AsyncMock())

        with (
            patch("check_active_v2.open_check_page", new=AsyncMock(return_value=True)),
            patch("check_active_v2.solve_captcha_task", new=AsyncMock(return_value=CaptchaSolution("ABCD", 1))),
            patch("check_active_v2.get_captcha_image_base64", new=AsyncMock(return_value="QUJDRA==")),
            patch(
                "check_active_v2.wait_for_result_payload",
                new=AsyncMock(return_value={"purchase_date": "01/03/2026"}),
            ),
            patch("check_active_v2.save_result_screenshot", new=AsyncMock()) as mock_screenshot,
        ):
            result = asyncio.run(check_serial(fake_page, "SN0000002", capture_screenshot=True, folder_name="batch"))

        self.assertEqual(result, ["SN0000002", "01/03/2026"])
        mock_screenshot.assert_awaited_once_with(fake_page, "SN0000002", "01/03/2026", "batch")



class SerialLoadingTests(unittest.TestCase):
    def test_normalize_serial_strips_bom_and_punctuation(self):
        self.assertEqual(normalize_serial("\ufeffc02x g2-jrd"), "C02XG2JRD")

    def test_is_valid_serial_rejects_wrong_length(self):
        self.assertTrue(is_valid_serial("C02XG2JRD"))
        self.assertFalse(is_valid_serial("SHORT"))
        self.assertFalse(is_valid_serial("WAYTOOLONGSERIAL"))

    def test_load_serials_handles_bom_and_removes_duplicates(self):
        with NamedTemporaryFile("w", suffix=".txt", delete=False, encoding="utf-8-sig") as handle:
            handle.write("C02XG2JRD\nc02xg2jrd\n\nD96XVQJRD8\n")
            path = handle.name

        try:
            self.assertEqual(load_serials(path), ["C02XG2JRD", "D96XVQJRD8"])
        finally:
            os.unlink(path)

    def test_needs_check_reruns_only_failed_results(self):
        done = {"AAA11111": "01/03/2026", "BBB22222": "Check tay", "CCC33333": "Chưa active"}
        self.assertFalse(needs_check("AAA11111", done))
        self.assertTrue(needs_check("BBB22222", done))
        self.assertFalse(needs_check("CCC33333", done))
        self.assertTrue(needs_check("DDD44444", done))


class CaptchaTimingTests(unittest.TestCase):
    def test_2captcha_polling_is_fast_for_interactive_app(self):
        # Hỏi sớm và hỏi dày: bắt được lúc 2captcha vừa xong thay vì ngủ đủ 2s.
        self.assertLessEqual(CAPTCHA_2CAPTCHA_FIRST_POLL_DELAY_SECONDS, 1)
        self.assertLessEqual(CAPTCHA_2CAPTCHA_POLL_INTERVAL_SECONDS, 0.5)

    def test_2captcha_total_wait_ceiling_did_not_shrink(self):
        # Hỏi dày hơn mà không tăng số vòng = bỏ mã còn đang giải dở, mất tiền.
        ceiling = (
            CAPTCHA_2CAPTCHA_FIRST_POLL_DELAY_SECONDS
            + CAPTCHA_2CAPTCHA_MAX_POLLS * CAPTCHA_2CAPTCHA_POLL_INTERVAL_SECONDS
        )
        self.assertGreaterEqual(ceiling, 20)


class ProxyTests(unittest.TestCase):
    def test_parse_proxy_line_reads_url_format_with_credentials(self):
        self.assertEqual(
            parse_proxy_line("http://user:p%40ss@gate.example.com:8000"),
            {"server": "http://gate.example.com:8000", "username": "user", "password": "p@ss"},
        )

    def test_parse_proxy_line_reads_host_port_user_pass_format(self):
        self.assertEqual(
            parse_proxy_line("1.2.3.4:8000:user:pass"),
            {"server": "http://1.2.3.4:8000", "username": "user", "password": "pass"},
        )

    def test_parse_proxy_line_skips_comments_and_blank_lines(self):
        self.assertIsNone(parse_proxy_line("# ghi chú"))
        self.assertIsNone(parse_proxy_line("   "))

    def test_proxy_for_index_cycles_and_handles_empty_list(self):
        proxies = [{"server": "a"}, {"server": "b"}]
        self.assertEqual(proxy_for_index(proxies, 0)["server"], "a")
        self.assertEqual(proxy_for_index(proxies, 3)["server"], "b")
        self.assertIsNone(proxy_for_index([], 0))

    def test_load_proxies_returns_empty_list_when_file_missing(self):
        self.assertEqual(load_proxies(Path("khong-ton-tai-proxies.txt")), [])


class BlockHandlingTests(unittest.IsolatedAsyncioTestCase):
    def test_block_throttle_backs_off_exponentially_and_resets(self):
        throttle = BlockThrottle(base_delay=10, max_delay=40)
        self.assertEqual(throttle.next_delay(), 10)
        throttle.block_count = 2
        self.assertEqual(throttle.next_delay(), 20)
        throttle.block_count = 9
        self.assertEqual(throttle.next_delay(), 40)
        throttle.on_success()
        self.assertEqual(throttle.block_count, 0)

    async def test_open_check_page_raises_blocked_error_on_403(self):
        class FakeResponse:
            status = 403
            headers = {"retry-after": "120"}

        class FakePage:
            async def goto(self, *_args, **_kwargs):
                return FakeResponse()

        with self.assertRaises(BlockedError) as ctx:
            await open_check_page(FakePage(), "C02XG2JRD")

        self.assertIn("403", str(ctx.exception))

    async def test_open_check_page_raises_blocked_error_on_dead_proxy(self):
        class FakePage:
            async def goto(self, *_args, **_kwargs):
                raise RuntimeError("net::ERR_TUNNEL_CONNECTION_FAILED at https://...")

        with self.assertRaises(BlockedError):
            await open_check_page(FakePage(), "C02XG2JRD")

    async def test_open_check_page_does_not_wait_for_network_idle_before_typing(self):
        class FakePage:
            def __init__(self):
                self.waited_for_network_idle = False
                self.value = ""

            async def goto(self, *_args, **_kwargs):
                class FakeResponse:
                    status = 200

                return FakeResponse()

            async def wait_for_load_state(self, *_args, **_kwargs):
                self.waited_for_network_idle = True

            async def fill(self, selector, value):
                self.filled = (selector, value)
                self.value = value

            async def wait_for_timeout(self, *_args, **_kwargs):
                return None

            async def input_value(self, selector):
                return self.value

        page = FakePage()
        self.assertTrue(await open_check_page(page, "C02XG2JRD"))
        self.assertFalse(page.waited_for_network_idle)
        self.assertEqual(page.filled, ("#serial-number-input", "C02XG2JRD"))

    async def test_fill_serial_number_retries_when_page_hydration_clears_value(self):
        class FakePage:
            def __init__(self):
                self.fill_count = 0
                self.value = ""

            async def fill(self, selector, value):
                self.fill_count += 1
                self.value = value

            async def wait_for_timeout(self, *_args, **_kwargs):
                if self.fill_count == 1:
                    self.value = ""

            async def input_value(self, selector):
                return self.value

        page = FakePage()
        self.assertTrue(await fill_serial_number(page, "C02XG2JRD"))
        self.assertEqual(page.fill_count, 2)



class ArgumentTests(unittest.TestCase):
    def test_force_flag_defaults_to_false(self):
        self.assertFalse(parse_args([]).force)

    def test_force_flag_can_be_enabled(self):
        self.assertTrue(parse_args(["--force"]).force)



class StaleResultPageTests(unittest.IsolatedAsyncioTestCase):
    """Context dùng lại cho nhiều serial thì phải xoá cookie, nếu không sẽ
    vào lại đúng trang kết quả của serial trước."""

    def _fake_page(self, has_serial_input=True):
        calls = {"clear_cookies": 0, "goto": 0, "fill": 0, "evaluate": 0}

        class FakeContext:
            async def clear_cookies(self):
                calls["clear_cookies"] += 1

        class FakeResponse:
            status = 200
            headers = {}

        class FakePage:
            context = FakeContext()

            async def goto(self, *_args, **_kwargs):
                calls["goto"] += 1
                return FakeResponse()

            async def evaluate(self, *_args, **_kwargs):
                calls["evaluate"] += 1

            async def wait_for_load_state(self, *_args, **_kwargs):
                return None

            async def wait_for_timeout(self, *_args, **_kwargs):
                return None

            async def fill(self, *_args, **_kwargs):
                calls["fill"] += 1
                if not has_serial_input:
                    raise RuntimeError(
                        "Page.fill: Timeout 30000ms exceeded waiting for locator(\"#serial-number-input\")"
                    )
                self.value = _args[1] if len(_args) > 1 else ""

            async def input_value(self, *_args, **_kwargs):
                return getattr(self, "value", "")

        return FakePage(), calls

    async def test_clear_site_state_clears_cookies_and_storage(self):
        page, calls = self._fake_page()
        await clear_site_state(page)
        self.assertEqual(calls["clear_cookies"], 1)
        self.assertEqual(calls["evaluate"], 1)

    async def test_open_check_page_clears_cookies_before_navigating(self):
        page, calls = self._fake_page(has_serial_input=True)
        self.assertTrue(await open_check_page(page, "C02XG2JRD"))
        self.assertEqual(calls["clear_cookies"], 1)
        self.assertEqual(calls["goto"], 1)

    async def test_open_check_page_reclears_when_serial_input_missing(self):
        page, calls = self._fake_page(has_serial_input=False)
        self.assertFalse(await open_check_page(page, "C02XG2JRD"))
        # 1 lan truoc khi vao + 2 lan retry khi khong thay o nhap serial
        self.assertEqual(calls["clear_cookies"], 3)
        self.assertEqual(calls["goto"], 3)



class AssetBlockingTests(unittest.IsolatedAsyncioTestCase):
    """Chặn ảnh/font để đỡ tốn băng thông proxy."""

    def _fake_route(self, resource_type, url="https://cdn-apple.com/asset"):
        calls = {"abort": 0, "continue": 0}

        class FakeRequest:
            pass

        class FakeRoute:
            def __init__(self):
                self.request = FakeRequest()
                self.request.resource_type = resource_type
                self.request.url = url

            async def abort(self):
                calls["abort"] += 1

            async def continue_(self):
                calls["continue"] += 1

        return FakeRoute(), calls

    async def test_images_and_fonts_are_aborted(self):
        for resource_type in ("image", "font", "media"):
            route, calls = self._fake_route(resource_type)
            await block_heavy_assets(route)
            self.assertEqual(calls["abort"], 1, resource_type)
            self.assertEqual(calls["continue"], 0, resource_type)

    async def test_document_and_script_pass_through(self):
        for resource_type in ("document", "script", "xhr", "stylesheet", "fetch"):
            route, calls = self._fake_route(resource_type)
            await block_heavy_assets(route)
            self.assertEqual(calls["abort"], 0, resource_type)
            self.assertEqual(calls["continue"], 1, resource_type)

    async def test_captcha_image_from_check_page_passes_through(self):
        route, calls = self._fake_route("image", "https://checkcoverage.apple.com/captcha.png")
        await block_heavy_assets(route)
        self.assertEqual(calls["abort"], 0)
        self.assertEqual(calls["continue"], 1)

    def test_blocked_types_never_include_document_or_script(self):
        self.assertNotIn("document", BLOCKED_RESOURCE_TYPES)
        self.assertNotIn("script", BLOCKED_RESOURCE_TYPES)
        self.assertNotIn("xhr", BLOCKED_RESOURCE_TYPES)



class ProxyFailureTests(unittest.IsolatedAsyncioTestCase):
    """Proxy chậm/chết KHÔNG được làm sập cả run."""

    class FakePage:
        def __init__(self, goto_error=None):
            self.goto_error = goto_error

        async def goto(self, *_args, **_kwargs):
            if self.goto_error:
                raise self.goto_error
            raise AssertionError("không nên tới đây")

        async def evaluate(self, *_args, **_kwargs):
            return None

        async def wait_for_timeout(self, *_args, **_kwargs):
            return None

    async def test_goto_timeout_becomes_proxy_failure(self):
        """Day chinh la loi lam sap run truoc do."""
        error = Exception("Page.goto: Timeout 30000ms exceeded.\nCall log:\n  - navigating to ...")
        with self.assertRaises(ProxyFailure):
            await open_check_page(self.FakePage(error), "C02XG2JRD")

    async def test_dns_failure_becomes_proxy_failure(self):
        error = Exception("net::ERR_NAME_NOT_RESOLVED at https://checkcoverage.apple.com")
        with self.assertRaises(ProxyFailure):
            await open_check_page(self.FakePage(error), "C02XG2JRD")

    async def test_closed_page_is_not_swallowed_as_proxy_failure(self):
        """Dang tat chuong trinh thi phai thoat han, khong doi proxy vo nghia."""
        error = Exception("Target page, context or browser has been closed")
        with self.assertRaises(Exception) as ctx:
            await open_check_page(self.FakePage(error), "C02XG2JRD")
        self.assertNotIsInstance(ctx.exception, ProxyFailure)

    def test_proxy_failure_is_still_a_blocked_error(self):
        """Code cu bat BlockedError van phai bat duoc ProxyFailure."""
        self.assertTrue(issubclass(ProxyFailure, BlockedError))


class PageClosedDetectionTests(unittest.TestCase):
    def test_detects_playwright_closed_messages(self):
        self.assertTrue(is_page_closed_error(Exception("Target page, context or browser has been closed")))
        self.assertTrue(is_page_closed_error(Exception("Locator.count: Target closed")))

    def test_ordinary_errors_are_not_closed_errors(self):
        self.assertFalse(is_page_closed_error(Exception("Timeout 30000ms exceeded")))
        self.assertFalse(is_page_closed_error(Exception("net::ERR_TUNNEL_CONNECTION_FAILED")))


class DeadProxyTests(unittest.TestCase):
    def setUp(self):
        self.proxies = [{"server": "http://a:1"}, {"server": "http://b:2"}, {"server": "http://c:3"}]

    def test_cycles_through_all_when_none_dead(self):
        picked = [pick_proxy(self.proxies, i)["server"] for i in range(4)]
        self.assertEqual(picked, ["http://a:1", "http://b:2", "http://c:3", "http://a:1"])

    def test_skips_dead_proxies(self):
        dead = {"http://b:2"}
        picked = {pick_proxy(self.proxies, i, dead)["server"] for i in range(6)}
        self.assertEqual(picked, {"http://a:1", "http://c:3"})

    def test_returns_none_when_all_proxies_dead(self):
        dead = {"http://a:1", "http://b:2", "http://c:3"}
        self.assertIsNone(pick_proxy(self.proxies, 0, dead))

    def test_returns_none_when_no_proxies_configured(self):
        self.assertIsNone(pick_proxy([], 0))



class CaptchaImageNotBlockedTests(unittest.TestCase):
    """Chan anh de tiet kiem bang thong, nhung KHONG duoc chan anh captcha."""

    def test_cdn_images_are_blocked(self):
        self.assertTrue(should_block_request("image", "https://store.storeimages.cdn-apple.com/hero.jpg"))
        self.assertTrue(should_block_request("font", "https://www.apple.com/wss/fonts/SF-Pro.woff2"))
        self.assertTrue(should_block_request("media", "https://apple.com/video.mp4"))

    def test_images_from_the_check_page_are_never_blocked(self):
        self.assertFalse(
            should_block_request("image", "https://checkcoverage.apple.com/assets/captcha/abc.png")
        )

    def test_page_and_scripts_always_pass(self):
        for resource_type in ("document", "script", "xhr", "fetch", "stylesheet"):
            self.assertFalse(should_block_request(resource_type, "https://checkcoverage.apple.com/x"))



class CaptchaLengthTests(unittest.IsolatedAsyncioTestCase):
    """Da tra tien cho mot ma roi thi phai GUI THU, dung vut di."""

    def _fake_page(self):
        locator = AsyncMock()
        locator.first = locator

        class FakePage:
            def __init__(self, locator, keyboard):
                self._locator = locator
                self.keyboard = keyboard

            def locator(self, _selector):
                return self._locator

        return FakePage(locator, AsyncMock())

    async def _check_with_code(self, code):
        reloads = {"count": 0}

        async def fake_reload(_page):
            reloads["count"] += 1

        with (
            patch("check_active_v2.open_check_page", new=AsyncMock(return_value=True)),
            patch("check_active_v2.has_invalid_serial_input_error", new=AsyncMock(return_value=False)),
            patch("check_active_v2.get_captcha_image_base64", new=AsyncMock(return_value="QUJDRA==")),
            patch("check_active_v2.solve_captcha_task", new=AsyncMock(return_value=CaptchaSolution(code, 1))),
            patch("check_active_v2.wait_for_result_payload",
                  new=AsyncMock(return_value={"purchase_date": "01/03/2026"})),
            patch("check_active_v2.reload_captcha", new=fake_reload),
            patch("check_active_v2.report_bad_captcha", new=AsyncMock()),
        ):
            result = await check_serial(self._fake_page(), "C02XG2JRD", capture_screenshot=False)
        return result, reloads["count"]

    async def test_valid_serial_goes_straight_to_captcha_without_extra_error_check(self):
        page = self._fake_page()
        invalid_check = AsyncMock(return_value=False)

        with (
            patch("check_active_v2.open_check_page", new=AsyncMock(return_value=True)),
            patch("check_active_v2.has_invalid_serial_input_error", new=invalid_check),
            patch("check_active_v2.get_captcha_image_base64", new=AsyncMock(return_value="QUJDRA==")),
            patch("check_active_v2.solve_captcha_task", new=AsyncMock(return_value=CaptchaSolution("ABCD", 1))),
            patch("check_active_v2.wait_for_result_payload",
                  new=AsyncMock(return_value={"purchase_date": "01/03/2026"})),
        ):
            result = await check_serial(page, "C02XG2JRD", capture_screenshot=False)

        self.assertEqual(result, ["C02XG2JRD", "01/03/2026"])
        invalid_check.assert_not_awaited()

    async def test_five_character_code_is_submitted_not_discarded(self):
        """Truoc day 42% ma tra tien bi vut o day."""
        result, reloads = await self._check_with_code("ABCDE")
        self.assertEqual(result, ["C02XG2JRD", "01/03/2026"])
        self.assertEqual(reloads, 0, "không được reload, phải gửi thử")

    async def test_four_character_code_still_works(self):
        result, reloads = await self._check_with_code("ABCD")
        self.assertEqual(result, ["C02XG2JRD", "01/03/2026"])
        self.assertEqual(reloads, 0)

    async def test_obvious_garbage_is_still_rejected(self):
        result, reloads = await self._check_with_code("A")
        self.assertEqual(result, ["C02XG2JRD", "Check tay"])
        self.assertEqual(reloads, MAX_AUTO_RETRIES, "mã rác thì vẫn phải đổi mã khác")

    def test_accepted_length_range_covers_observed_answers(self):
        self.assertLessEqual(CAPTCHA_MIN_LENGTH, 4)
        self.assertGreaterEqual(CAPTCHA_MAX_LENGTH, 5)


class CaptchaSubmitTests(unittest.IsolatedAsyncioTestCase):
    async def test_submit_captcha_types_change_event_and_clicks_send(self):
        events = []

        class FakeInput:
            async def click(self):
                events.append("focus")

            async def clear(self):
                events.append("clear")

            async def press_sequentially(self, code, delay=None):
                events.append(("type", code, delay))

            async def evaluate(self, script):
                events.append(("event_script", script))

        class FakeButton:
            first = None

            def __init__(self):
                self.first = self

            async def click(self, timeout=None):
                events.append(("click_send", timeout))

        class FakePage:
            keyboard = AsyncMock()

            def __init__(self):
                self.input = FakeInput()
                self.button = FakeButton()

            def locator(self, selector):
                self.assertEqual(selector, '#captcha-input')
                return self.input

            def get_by_role(self, role, name=None):
                events.append(("find_button", role, bool(name)))
                return self.button

            def assertEqual(self, *args):
                testcase.assertEqual(*args)

        testcase = self
        page = FakePage()
        await submit_captcha_code(page, "HAL3")

        self.assertIn(("type", "HAL3", 20), events)
        event_scripts = [item[1] for item in events if isinstance(item, tuple) and item[0] == "event_script"]
        self.assertTrue(event_scripts)
        self.assertIn("change", event_scripts[0])
        self.assertIn(("click_send", 3000), events)
        page.keyboard.press.assert_not_awaited()


class CaptchaMismatchFlowTests(unittest.IsolatedAsyncioTestCase):
    def _fake_page(self):
        locator = AsyncMock()
        locator.first = locator

        class FakePage:
            keyboard = AsyncMock()

            def locator(self, _selector):
                return locator

        return FakePage()

    async def test_check_serial_reloads_immediately_when_apple_says_captcha_mismatch(self):
        payloads = [
            {"purchase_date": None, "captcha_error": True},
            {"purchase_date": "01/03/2026"},
        ]

        with (
            patch("check_active_v2.open_check_page", new=AsyncMock(return_value=True)),
            patch("check_active_v2.grab_captcha_image", new=AsyncMock(return_value="QUJDRA==")),
            patch("check_active_v2.solve_captcha_task",
                  new=AsyncMock(return_value=CaptchaSolution("HGEL", 99))),
            patch("check_active_v2.submit_captcha_code", new=AsyncMock()),
            patch("check_active_v2.wait_for_result_payload", new=AsyncMock(side_effect=payloads)),
            patch("check_active_v2.report_bad_captcha", new=AsyncMock()) as report_bad,
            patch("check_active_v2.reload_captcha", new=AsyncMock()) as reload_mock,
        ):
            result = await check_serial(self._fake_page(), "C02XG2JRD", capture_screenshot=False)

        self.assertEqual(result, ["C02XG2JRD", "01/03/2026"])
        report_bad.assert_awaited_once_with(99)
        reload_mock.assert_awaited_once()


class ManualCaptchaTests(unittest.IsolatedAsyncioTestCase):
    """Serial nao OCR mai khong ra thi hoi nguoi dung nhap tay."""

    def setUp(self):
        import check_active_v2

        self.module = check_active_v2
        self._handler = check_active_v2.MANUAL_CAPTCHA_HANDLER

    def tearDown(self):
        self.module.MANUAL_CAPTCHA_HANDLER = self._handler

    def _fake_page(self):
        locator = AsyncMock()
        locator.first = locator

        class FakePage:
            def __init__(self, locator, keyboard):
                self._locator = locator
                self.keyboard = keyboard

            def locator(self, _selector):
                return self._locator

        return FakePage(locator, AsyncMock())

    async def test_handler_is_asked_only_after_ocr_keeps_failing(self):
        asked = []

        async def handler(image_base64, serial):
            asked.append(serial)
            return "MANU"

        self.module.MANUAL_CAPTCHA_HANDLER = handler
        payloads = [{"purchase_date": None}, {"purchase_date": None}, {"purchase_date": "01/03/2026"}]

        with (
            patch("check_active_v2.open_check_page", new=AsyncMock(return_value=True)),
            patch("check_active_v2.has_invalid_serial_input_error", new=AsyncMock(return_value=False)),
            patch("check_active_v2.get_captcha_image_base64", new=AsyncMock(return_value="QUJDRA==")),
            patch("check_active_v2.solve_captcha_task", new=AsyncMock(return_value=CaptchaSolution("ABCD", 1))),
            patch("check_active_v2.wait_for_result_payload", new=AsyncMock(side_effect=payloads)),
            patch("check_active_v2.has_error_message", new=AsyncMock(return_value=True)),
            patch("check_active_v2.reload_captcha", new=AsyncMock()),
            patch("check_active_v2.report_bad_captcha", new=AsyncMock()),
        ):
            result = await check_serial(self._fake_page(), "C02XG2JRD", capture_screenshot=False)

        self.assertEqual(result, ["C02XG2JRD", "01/03/2026"])
        self.assertEqual(asked, ["C02XG2JRD"], "chỉ hỏi 1 lần, ở lần thử thứ 3")

    async def test_no_handler_means_business_as_usual(self):
        self.module.MANUAL_CAPTCHA_HANDLER = None
        with (
            patch("check_active_v2.open_check_page", new=AsyncMock(return_value=True)),
            patch("check_active_v2.has_invalid_serial_input_error", new=AsyncMock(return_value=False)),
            patch("check_active_v2.get_captcha_image_base64", new=AsyncMock(return_value="QUJDRA==")),
            patch("check_active_v2.solve_captcha_task",
                  new=AsyncMock(return_value=CaptchaSolution("ABCD", 1))) as solver,
            patch("check_active_v2.wait_for_result_payload",
                  new=AsyncMock(return_value={"purchase_date": "01/03/2026"})),
        ):
            result = await check_serial(self._fake_page(), "C02XG2JRD", capture_screenshot=False)
        self.assertEqual(result, ["C02XG2JRD", "01/03/2026"])
        solver.assert_awaited()

    async def test_handler_returning_nothing_falls_back_to_2captcha(self):
        async def handler(image_base64, serial):
            return None  # nguoi dung bam Bo qua hoac het gio

        self.module.MANUAL_CAPTCHA_HANDLER = handler
        payloads = [{"purchase_date": None}, {"purchase_date": None}, {"purchase_date": "01/03/2026"}]

        with (
            patch("check_active_v2.open_check_page", new=AsyncMock(return_value=True)),
            patch("check_active_v2.has_invalid_serial_input_error", new=AsyncMock(return_value=False)),
            patch("check_active_v2.get_captcha_image_base64", new=AsyncMock(return_value="QUJDRA==")),
            patch("check_active_v2.solve_captcha_task",
                  new=AsyncMock(return_value=CaptchaSolution("ABCD", 1))) as solver,
            patch("check_active_v2.wait_for_result_payload", new=AsyncMock(side_effect=payloads)),
            patch("check_active_v2.has_error_message", new=AsyncMock(return_value=True)),
            patch("check_active_v2.reload_captcha", new=AsyncMock()),
            patch("check_active_v2.report_bad_captcha", new=AsyncMock()),
        ):
            result = await check_serial(self._fake_page(), "C02XG2JRD", capture_screenshot=False)

        self.assertEqual(result, ["C02XG2JRD", "01/03/2026"])
        self.assertEqual(solver.await_count, 3)

    async def test_handler_is_asked_at_most_once_even_when_ocr_never_lands(self):
        """Hỏi tay 3 lần liên tiếp là cách nhanh nhất chạm trần SERIAL_TIMEOUT."""
        asked = []

        async def handler(image_base64, serial):
            asked.append(serial)
            return None  # het gio, khong ai go

        self.module.MANUAL_CAPTCHA_HANDLER = handler

        with (
            patch("check_active_v2.open_check_page", new=AsyncMock(return_value=True)),
            patch("check_active_v2.has_invalid_serial_input_error", new=AsyncMock(return_value=False)),
            patch("check_active_v2.get_captcha_image_base64", new=AsyncMock(return_value="QUJDRA==")),
            patch("check_active_v2.solve_captcha_task",
                  new=AsyncMock(return_value=CaptchaSolution("ABCD", 1))),
            patch("check_active_v2.wait_for_result_payload",
                  new=AsyncMock(return_value={"purchase_date": None})),
            patch("check_active_v2.has_error_message", new=AsyncMock(return_value=True)),
            patch("check_active_v2.reload_captcha", new=AsyncMock()),
            patch("check_active_v2.report_bad_captcha", new=AsyncMock()),
        ):
            result = await check_serial(self._fake_page(), "C02XG2JRD", capture_screenshot=False)

        self.assertEqual(result, ["C02XG2JRD", "Check tay"])
        self.assertEqual(len(asked), 1, "5 lần thử nhưng chỉ được hỏi tay 1 lần")


class CaptchaPrefetchTests(unittest.IsolatedAsyncioTestCase):
    """Ảnh captcha phải được tải song song với lúc điền serial."""

    def _fake_page(self):
        locator = AsyncMock()
        locator.first = locator

        class FakePage:
            def __init__(self, locator):
                self._locator = locator
                self.keyboard = AsyncMock()

            def locator(self, _selector):
                return self._locator

        return FakePage(locator)

    async def test_captcha_grab_starts_before_serial_is_filled(self):
        order = []

        async def fake_open(page, sn, on_page_ready=None):
            order.append("trang tải xong")
            if on_page_ready is not None:
                on_page_ready()
            await asyncio.sleep(0)  # nhuong luot cho task prefetch chay
            order.append("điền serial xong")
            return True

        async def fake_grab(page):
            order.append("bắt đầu lấy ảnh captcha")
            return "QUJDRA=="

        with (
            patch("check_active_v2.open_check_page", new=fake_open),
            patch("check_active_v2.grab_captcha_image", new=fake_grab),
            patch("check_active_v2.solve_captcha_task",
                  new=AsyncMock(return_value=CaptchaSolution("ABCD", 1))),
            patch("check_active_v2.wait_for_result_payload",
                  new=AsyncMock(return_value={"purchase_date": "01/03/2026"})),
        ):
            result = await check_serial(self._fake_page(), "C02XG2JRD", capture_screenshot=False)

        self.assertEqual(result, ["C02XG2JRD", "01/03/2026"])
        self.assertEqual(
            order,
            ["trang tải xong", "bắt đầu lấy ảnh captcha", "điền serial xong"],
            "lấy ảnh captcha phải chạy chồng lên lúc điền serial",
        )

    async def test_prefetch_task_is_cancelled_when_page_fails_to_open(self):
        started = asyncio.Event()

        async def fake_open(page, sn, on_page_ready=None):
            if on_page_ready is not None:
                on_page_ready()
            await started.wait()
            return False

        async def never_ending(page):
            started.set()
            await asyncio.sleep(3600)

        with (
            patch("check_active_v2.open_check_page", new=fake_open),
            patch("check_active_v2.grab_captcha_image", new=never_ending),
        ):
            result = await asyncio.wait_for(
                check_serial(self._fake_page(), "C02XG2JRD", capture_screenshot=False),
                timeout=2,
            )

        self.assertEqual(result, ["C02XG2JRD", "Lỗi load trang"])


class SerialTaggedOutputTests(unittest.IsolatedAsyncioTestCase):
    """Chạy 2 luồng thì stdout dùng chung — mỗi dòng phải biết mình của serial nào."""

    def _fake_page(self):
        locator = AsyncMock()
        locator.first = locator

        class FakePage:
            def __init__(self):
                self.keyboard = AsyncMock()

            def locator(self, _selector):
                return locator

        return FakePage()

    async def test_two_concurrent_checks_tag_their_own_lines(self):
        import contextlib

        from check_active_v2 import CURRENT_SERIAL

        tagged = []

        class TaggingStdout:
            """Đúng cách app đọc nhãn: lấy ContextVar ngay lúc ghi."""

            def write(self, text):
                for line in text.splitlines():
                    if line.strip():
                        tagged.append((CURRENT_SERIAL.get(), line.strip()))
                return len(text)

            def flush(self):
                pass

        async def slow_result(_page):
            await asyncio.sleep(0.02)
            return {"purchase_date": "01/03/2026"}

        with (
            patch("check_active_v2.open_check_page", new=AsyncMock(return_value=True)),
            patch("check_active_v2.grab_captcha_image", new=AsyncMock(return_value="QUJDRA==")),
            patch("check_active_v2.solve_captcha_task",
                  new=AsyncMock(return_value=CaptchaSolution("ABCD", 1))),
            patch("check_active_v2.wait_for_result_payload", new=slow_result),
            contextlib.redirect_stdout(TaggingStdout()),
        ):
            await asyncio.gather(
                check_serial(self._fake_page(), "C02XG2JRD", capture_screenshot=False),
                check_serial(self._fake_page(), "DKVQL1WXYZ", capture_screenshot=False),
            )

        # Moi dong phai co nhan, va nhan phai khop voi serial trong dong do
        self.assertTrue(tagged, "khong bat duoc dong nao")
        for serial, line in tagged:
            self.assertIsNotNone(serial, f"dong khong co nhan: {line}")
        for serial in ("C02XG2JRD", "DKVQL1WXYZ"):
            lines = [line for tag, line in tagged if tag == serial]
            self.assertIn(f"🔍 Check: {serial}", lines)
            self.assertTrue(any("THÀNH CÔNG" in line for line in lines),
                            f"{serial} thieu dong ket qua")


class ReadResultTextsTests(unittest.IsolatedAsyncioTestCase):
    """Đọc 7 vùng text bằng MỘT lượt evaluate thay vì 14 lượt locator."""

    async def test_single_evaluate_result_is_used(self):
        page = AsyncMock()
        page.evaluate = AsyncMock(return_value={
            "device_title": " iPhone 15  ",
            "purchase_text": "Đã mua 01/03/2026",
            "notification_text": "",
            "header_text": "Số Sê-ri",
            "body_text": "toàn trang",
            "heading_text": "",
            "error_text": "",
        })

        payload = await read_result_texts(page)

        page.evaluate.assert_awaited_once()
        self.assertEqual(payload["device_title"], "iPhone 15")
        self.assertEqual(payload["purchase_text"], "Đã mua 01/03/2026")
        self.assertEqual(set(payload), set(RESULT_TEXT_KEYS))

    async def test_falls_back_to_locators_when_evaluate_is_unavailable(self):
        page = AsyncMock()
        page.evaluate = AsyncMock(side_effect=RuntimeError("không chạy được JS"))
        fallback = AsyncMock(return_value={key: "" for key in RESULT_TEXT_KEYS})

        with patch("check_active_v2.read_result_texts_via_locators", new=fallback):
            payload = await read_result_texts(page)

        fallback.assert_awaited_once()
        self.assertEqual(set(payload), set(RESULT_TEXT_KEYS))


class ScreenshotTests(unittest.IsolatedAsyncioTestCase):
    async def test_screenshot_captures_viewport_not_whole_page(self):
        page = AsyncMock()
        with patch("check_active_v2.build_screenshot_path", return_value=Path("/tmp/x.png")):
            await save_result_screenshot(page, "C02XG2JRD", "01/03/2026")
        self.assertIs(page.screenshot.await_args.kwargs["full_page"], False)


class ReloadCaptchaTests(unittest.IsolatedAsyncioTestCase):
    """Đổi captcha xong phải chờ ảnh đổi thật, không ngủ cứng rồi đi tiếp."""

    def _page(self, srcs):
        locator = AsyncMock()
        locator.first = locator
        locator.get_attribute = AsyncMock(side_effect=srcs)

        class FakePage:
            def __init__(self):
                self.waits = []

            def locator(self, _selector):
                return locator

            async def wait_for_timeout(self, ms):
                self.waits.append(ms)

        return FakePage(), locator

    async def test_returns_as_soon_as_image_src_changes(self):
        page, locator = self._page(["cu", "cu", "moi"])
        await reload_captcha(page)
        locator.click.assert_awaited_once()
        # 2 vong x 60ms, khong phai ngu cung 250ms
        self.assertEqual(page.waits, [CAPTCHA_RELOAD_POLL_MS, CAPTCHA_RELOAD_POLL_MS])

    async def test_gives_up_waiting_after_the_ceiling(self):
        page, _ = self._page(["cu"] * 500)
        await reload_captcha(page)
        self.assertEqual(
            sum(page.waits), CAPTCHA_RELOAD_MAX_WAIT_MS,
            "không được chờ ảnh mới vô hạn",
        )


# Text that lay tu trang Apple ngay 26/08/2026 (anh chup man hinh cua Lam)
TRANG_KET_QUA_THAT = """AppleCare và bảo hành
iPhone 17 Pro Max
Số Sê-ri: CY2QLQ6XTJ
Đã mua 28 tháng 4, 2026
Phạm vi bảo hành của bạn
Bảo hành giới hạn
Hết hạn: 27 tháng 4, 2027"""

# Trang con dang hien form captcha — CO nhan "Nhap ma trong anh"
TRANG_FORM_CAPTCHA = """Kiểm tra phạm vi bảo hành
Nhập số sê-ri
Nhập mã trong ảnh
Làm mới mã
Tiếp tục"""


class CaptchaErrorDetectionTests(unittest.TestCase):
    """Nhan cua o nhap captcha KHONG duoc coi la loi sai ma.

    Bug 26/08: CAPTCHA_MISMATCH_TEXTS co "Nhap ma trong anh" va "Lam moi ma" —
    do la nhan cua chinh o nhap va nut doi ma, luon co tren trang. Lai con quet
    ca body_text. Ket qua: vua gui ma DUNG xong, trang chua kip chuyen, da vo
    doan "sai ma", bao 2captcha la sai roi mua ma khac. Nguoi dung nhin thay
    trang ket qua hien day du ma app van bao sai.
    """

    def test_captcha_form_labels_are_not_an_error(self):
        self.assertFalse(contains_captcha_error_text(TRANG_FORM_CAPTCHA))
        self.assertFalse(contains_captcha_error_text("Nhập mã trong ảnh"))
        self.assertFalse(contains_captcha_error_text("Làm mới mã"))

    def test_the_real_error_message_is_still_caught(self):
        self.assertTrue(contains_captcha_error_text(
            "Mã bạn đã nhập không khớp với hình ảnh."))

    def test_result_page_is_never_mistaken_for_an_error(self):
        self.assertFalse(contains_captcha_error_text(TRANG_KET_QUA_THAT))

    def test_empty_and_none_are_safe(self):
        self.assertFalse(contains_captcha_error_text(None, "", "   "))


class RealApplePageTests(unittest.TestCase):
    """Doc dung ngay mua tu trang that, dinh dang 'Da mua 28 thang 4, 2026'."""

    def test_purchase_date_is_read_from_the_real_page(self):
        ngay = determine_purchase_date(
            device_title="iPhone 17 Pro Max",
            header_text="Số Sê-ri: CY2QLQ6XTJ\nĐã mua 28 tháng 4, 2026",
            body_text=TRANG_KET_QUA_THAT,
        )
        self.assertEqual(ngay, "28/04/2026")

    def test_result_page_is_recognised_as_a_result(self):
        self.assertTrue(has_result_page_signal(
            "iPhone 17 Pro Max",
            "Số Sê-ri: CY2QLQ6XTJ\nĐã mua 28 tháng 4, 2026",
            body_text=TRANG_KET_QUA_THAT,
        ))


class WaitForResultPayloadTests(unittest.IsolatedAsyncioTestCase):
    """Trang dang chuyen tiep khong duoc ket luan la sai ma."""

    def _page(self, cac_lan_doc):
        class FakePage:
            def __init__(self):
                self.con_lai = list(cac_lan_doc)
                self.so_lan_cho = 0

            async def evaluate(self, _js):
                return self.con_lai.pop(0) if len(self.con_lai) > 1 else self.con_lai[0]

            async def wait_for_timeout(self, _ms):
                self.so_lan_cho += 1

        return FakePage()

    def _payload(self, **kwargs):
        base = {k: "" for k in RESULT_TEXT_KEYS}
        base.update(kwargs)
        return base

    async def test_waits_through_the_transition_instead_of_crying_wrong_code(self):
        """Lan doc dau con thay form, lan sau moi ra ket qua — phai cho."""
        page = self._page([
            self._payload(body_text=TRANG_FORM_CAPTCHA),
            self._payload(body_text=TRANG_FORM_CAPTCHA),
            self._payload(
                device_title="iPhone 17 Pro Max",
                header_text="Số Sê-ri: CY2QLQ6XTJ\nĐã mua 28 tháng 4, 2026",
                body_text=TRANG_KET_QUA_THAT,
            ),
        ])
        payload = await wait_for_result_payload(page)
        self.assertEqual(payload["purchase_date"], "28/04/2026")
        self.assertFalse(payload.get("captcha_error"),
                         "khong duoc bao sai ma khi ma that ra la dung")

    async def test_a_genuine_wrong_code_is_still_reported(self):
        page = self._page([self._payload(
            error_text="Mã bạn đã nhập không khớp với hình ảnh.",
            body_text=TRANG_FORM_CAPTCHA,
        )])
        payload = await wait_for_result_payload(page)
        self.assertTrue(payload.get("captcha_error"))
        self.assertIsNone(payload["purchase_date"])


class SerialClearedByPageTests(unittest.IsolatedAsyncioTestCase):
    """Trang Apple hydrate xong hay xoa trang o nhap serial.

    Anh chup 26/08: o serial TRONG, vien do, "Vui long nhap so se-ri hop le.",
    trong khi o CAPTCHA da dien dung "QW7T". Kiem tra mot lan luc dien la khong
    du — phai kiem lai ngay truoc khi bam gui.
    """

    def _page(self, cac_gia_tri):
        class FakePage:
            def __init__(self):
                self.doc = list(cac_gia_tri)
                self.da_dien = []

            async def input_value(self, _sel):
                return self.doc.pop(0) if len(self.doc) > 1 else self.doc[0]

            async def fill(self, _sel, value):
                self.da_dien.append(value)
                self.doc = [value]

            async def wait_for_timeout(self, _ms):
                pass

        return FakePage()

    async def test_refills_when_the_page_wiped_the_field(self):
        page = self._page([""])          # o da bi xoa trang
        ok = await ensure_serial_filled(page, "CY2QLQ6XTJ")
        self.assertTrue(ok)
        self.assertEqual(page.da_dien, ["CY2QLQ6XTJ"], "phai dien lai serial")

    async def test_does_nothing_when_the_field_is_still_correct(self):
        page = self._page(["CY2QLQ6XTJ"])
        ok = await ensure_serial_filled(page, "CY2QLQ6XTJ")
        self.assertTrue(ok)
        self.assertEqual(page.da_dien, [], "dang dung thi khong duoc dien lai")

    async def test_field_holding_a_different_serial_is_corrected(self):
        page = self._page(["HWV617QRXQ"])
        await ensure_serial_filled(page, "CY2QLQ6XTJ")
        self.assertEqual(page.da_dien, ["CY2QLQ6XTJ"])

    async def test_unreadable_field_does_not_block_the_run(self):
        class Hong:
            async def input_value(self, _sel):
                raise RuntimeError("khong doc duoc")

        self.assertTrue(await ensure_serial_filled(Hong(), "CY2QLQ6XTJ"))


class MissingSerialIsNotACaptchaErrorTests(unittest.IsolatedAsyncioTestCase):
    """Thieu serial thi KHONG duoc bao 2captcha la giai sai — vua mat tien vua do oan."""

    def _fake_page(self):
        locator = AsyncMock()
        locator.first = locator

        class FakePage:
            def __init__(self):
                self.keyboard = AsyncMock()

            def locator(self, _selector):
                return locator

            async def input_value(self, _sel):
                return "CY2QLQ6XTJ"

            async def fill(self, _sel, _v):
                return None

            async def wait_for_timeout(self, _ms):
                return None

        return FakePage()

    async def test_apple_asking_for_the_serial_does_not_blame_2captcha(self):
        bao_sai = AsyncMock()
        ket_qua = [
            {"purchase_date": None},
            {"purchase_date": "28/04/2026"},
        ]
        # Lan dau Apple bao thieu serial, lan sau thi xong
        thieu_serial = AsyncMock(side_effect=[True, False])

        with (
            patch("check_active_v2.open_check_page", new=AsyncMock(return_value=True)),
            patch("check_active_v2.grab_captcha_image", new=AsyncMock(return_value="QUJDRA==")),
            patch("check_active_v2.solve_captcha_task",
                  new=AsyncMock(return_value=CaptchaSolution("QW7T", 1))),
            patch("check_active_v2.submit_captcha_code", new=AsyncMock()),
            patch("check_active_v2.wait_for_result_payload",
                  new=AsyncMock(side_effect=ket_qua)),
            patch("check_active_v2.has_invalid_serial_input_error", new=thieu_serial),
            patch("check_active_v2.report_bad_captcha", new=bao_sai),
            patch("check_active_v2.reload_captcha", new=AsyncMock()),
            patch("check_active_v2.fill_serial_number", new=AsyncMock(return_value=True)),
        ):
            ket = await check_serial(self._fake_page(), "CY2QLQ6XTJ",
                                     capture_screenshot=False)

        self.assertEqual(ket, ["CY2QLQ6XTJ", "28/04/2026"])
        bao_sai.assert_not_awaited()


if __name__ == "__main__":
    unittest.main()
